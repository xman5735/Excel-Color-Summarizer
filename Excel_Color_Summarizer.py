import openpyxl
import tkinter as tk
from tkinter import filedialog
import math

def convert_cells_to_numbers(filename):
    # Load the workbook and get the active sheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Iterate through all the cells in the sheet for conversion
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # Check if the cell value is not None
                # Try converting the cell value to a float or integer
                try:
                    value = float(cell.value)
                    if value.is_integer():
                        cell.value = int(value)
                    else:
                        cell.value = value
                except ValueError:
                    # If the conversion fails, just continue to the next cell
                    pass

    # List to store column indices that need to be deleted
    columns_to_delete = []

    # Check the value in row 6 for each column
    for col_num, col_cells in enumerate(ws.iter_cols(), 1):
        if col_cells[5].value == 0:  # 5 is the index for row 6 (0-indexed)
            columns_to_delete.append(col_num)

    # Delete columns in reverse order (to avoid shifting issues)
    for col_num in reversed(columns_to_delete):
        ws.delete_cols(col_num)


    # Function to calculate average and standard deviation
    def get_stats(values):
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        std_dev = math.sqrt(variance)
        return mean, std_dev

    # Calculate and place averages and standard deviations
    for row_num, target_row in zip([6, 8, 10], [20, 21, 22]):
        values = [cell.value for cell in ws[row_num] if isinstance(cell.value, (int, float))]
        mean, std_dev = get_stats(values)

        ws[f'B{target_row}'] = mean
        ws[f'C{target_row}'] = std_dev
        ws[f'D{target_row}'] = 2 * std_dev

    # Place average ± deviations
    for row_num, target_row in zip([6, 8, 10], [25, 26, 27]):
        mean = ws[f'B{target_row - 5}'].value
        std_dev = ws[f'C{target_row - 5}'].value

        ws[f'B{target_row}'] = mean + std_dev
        ws[f'C{target_row}'] = mean - std_dev

    for row_num, target_row in zip([6, 8, 10], [30, 31, 32]):
        mean = ws[f'B{target_row - 10}'].value
        second_dev = ws[f'D{target_row - 10}'].value

        ws[f'B{target_row}'] = mean + second_dev
        ws[f'C{target_row}'] = mean - second_dev

    # Save the modified workbook
    wb.save(filename)

if __name__ == "__main__":
    # Create a simple tkinter window (but don't show it)
    root = tk.Tk()
    root.withdraw()

    # Open a file dialog asking the user to select an Excel file
    filename = filedialog.askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "*.xlsx")])
    
    # Check if a file was selected
    if filename:
        convert_cells_to_numbers(filename)
        print(f"Processed and saved: {filename}")
    else:
        print("No file selected.")
